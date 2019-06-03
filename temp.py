from peewee import SqliteDatabase, IntegerField, Model, DateTimeField, CharField, ForeignKeyField
from peewee import datetime as peewee_datetime
from models import User
from openpyxl import load_workbook

wb_base = load_workbook(filename='workbook1.xlsx', read_only=False, keep_vba=True)
ws_data = wb_base['sheet2']


for i in range(2, ws_data.max_row):
    
    User.create(
    
        user_id = ws_data.cell(row=i, column=8).value,
        username = ws_data.cell(row=i, column=5).value,
        first_name = ws_data.cell(row=i, column=6).value,
        last_name = ws_data.cell(row=i, column=7).value,
    
        house = 2,
        section = ws_data.cell(row=i, column=4).value,
        floor = ws_data.cell(row=i, column=3).value,
        apartment = ws_data.cell(row=i, column=9).value,
    )


# query = User.select().where(User.id > 10)
# # query = User.delete().where(User.id > 10)
# # query.execute()

# for i in query:
#     print(i.id)

# # print(peewee_datetime.datetime.now())
# # print(peewee_datetime.datetime.today().day)
