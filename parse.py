# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import os
import re
from openpyxl import load_workbook

wb = load_workbook('temp.xlsx')
sh = wb['Sheet1']

files_dir = os.path.join('chat', 'ChatExport_13_08_2019')
files = files_dir = [f for f in os.listdir(files_dir) if f.endswith('.html')]

for file in files:
    html = open(os.path.join('chat', 'ChatExport_13_08_2019', file), encoding="utf8").read()
    soup = BeautifulSoup(html, 'lxml')
    lastrow = sh.max_row

    msgs = soup.find_all('div', class_=re.compile(r'^message default clearfix'))
    for r, i in enumerate(msgs, start=lastrow):
        try:
            name = i.find('div', class_='from_name').string
        except AttributeError:
            name = sh.cell(row=r-1, column=1).value
        sh.cell(row=r, column=1).value = str(name)
        try:
            msg = i.find('div', class_='text').text
            sh.cell(row=r, column=2).value = str(msg)
        except AttributeError:
            print(i.get('id'))
        date = i.find('div', class_='pull_right date details').get('title')
        sh.cell(row=r, column=3).value = str(date)

wb.save('temp.xlsx')
